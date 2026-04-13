from datetime import datetime
import csv
import io
import json
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from db_utils import get_connection, get_table_columns, resolve_table_name


router = APIRouter()


class ExportRequest(BaseModel):
    filters: Optional[Dict[str, Any]] = None
    fields: Optional[List[str]] = None
    limit: int = 10000
    dataset: str = "default"


def quote_identifier(identifier: str) -> str:
    escaped = identifier.replace('"', '""')
    return f'"{escaped}"'


def normalize_export_request(
    request: Optional[ExportRequest],
    dataset: Optional[str],
    limit: Optional[int],
) -> ExportRequest:
    if request is None:
        return ExportRequest(
            dataset=dataset or "default",
            limit=limit or 10000,
        )

    payload = request.model_dump() if hasattr(request, "model_dump") else request.dict()
    if dataset is not None:
        payload["dataset"] = dataset
    if limit is not None:
        payload["limit"] = limit
    return ExportRequest(**payload)


def get_export_data(
    filters: Optional[Dict[str, Any]] = None,
    limit: int = 10000,
    dataset: str = "default",
    fields: Optional[List[str]] = None,
):
    try:
        table_name = resolve_table_name(dataset)
        valid_columns = set(get_table_columns(table_name))

        selected_fields = [field for field in (fields or []) if field in valid_columns]
        select_clause = "*"
        if selected_fields:
            select_clause = ", ".join(quote_identifier(field) for field in selected_fields)

        with get_connection() as conn:
            cursor = conn.cursor()
            sql = f'SELECT {select_clause} FROM {quote_identifier(table_name)}'
            params = []

            if filters:
                conditions = []
                for key, value in filters.items():
                    if key not in valid_columns or value is None:
                        continue
                    if isinstance(value, (list, tuple)):
                        normalized_values = [item for item in value if item is not None]
                        if not normalized_values:
                            continue
                        placeholders = ",".join(["?" for _ in normalized_values])
                        conditions.append(f"{quote_identifier(key)} IN ({placeholders})")
                        params.extend(normalized_values)
                    else:
                        conditions.append(f"{quote_identifier(key)} = ?")
                        params.append(value)
                if conditions:
                    sql += " WHERE " + " AND ".join(conditions)

            sql += " LIMIT ?"
            params.append(limit)
            cursor.execute(sql, params)
            rows = cursor.fetchall()
            columns = [description[0] for description in cursor.description]
            return [dict(zip(columns, row)) for row in rows]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get data: {str(e)}")


def build_filename(dataset: str, extension: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{dataset}_nonprofits_export_{timestamp}.{extension}"


@router.post("/export/csv")
async def export_csv(
    request: Optional[ExportRequest] = Body(None),
    limit: Optional[int] = Query(None),
    dataset: Optional[str] = Query(None),
):
    try:
        export_request = normalize_export_request(request, dataset, limit)
        data = get_export_data(
            export_request.filters,
            export_request.limit,
            export_request.dataset,
            export_request.fields,
        )
        if not data:
            raise HTTPException(status_code=404, detail="No data found")

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        csv_bytes = io.BytesIO(output.getvalue().encode("utf-8"))
        filename = build_filename(export_request.dataset, "csv")
        return StreamingResponse(
            csv_bytes,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/export/json")
async def export_json(
    request: Optional[ExportRequest] = Body(None),
    limit: Optional[int] = Query(None),
    dataset: Optional[str] = Query(None),
):
    try:
        export_request = normalize_export_request(request, dataset, limit)
        data = get_export_data(
            export_request.filters,
            export_request.limit,
            export_request.dataset,
            export_request.fields,
        )
        if not data:
            raise HTTPException(status_code=404, detail="No data found")

        filename = build_filename(export_request.dataset, "json")
        return Response(
            content=json.dumps(data, indent=2, ensure_ascii=False),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/export/excel")
async def export_excel(
    request: Optional[ExportRequest] = Body(None),
    limit: Optional[int] = Query(None),
    dataset: Optional[str] = Query(None),
):
    try:
        try:
            from openpyxl import Workbook
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel export requires openpyxl. Install with: pip install openpyxl")

        export_request = normalize_export_request(request, dataset, limit)
        data = get_export_data(
            export_request.filters,
            export_request.limit,
            export_request.dataset,
            export_request.fields,
        )
        if not data:
            raise HTTPException(status_code=404, detail="No data found")

        wb = Workbook()
        ws = wb.active
        ws.title = "ProPublica Data"

        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row_num, record in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=record.get(header, ""))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = build_filename(export_request.dataset, "xlsx")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/status")
async def export_status(dataset: str = "default"):
    try:
        table_name = resolve_table_name(dataset)
        available_columns = set(get_table_columns(table_name))

        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(f'SELECT COUNT(*) FROM "{table_name}"')
            total_count = cursor.fetchone()[0]

            income_stats = (None, None, None)
            if "part_i_summary_12_total_revenue_cy" in available_columns:
                cursor.execute(
                    f'SELECT MIN(part_i_summary_12_total_revenue_cy), MAX(part_i_summary_12_total_revenue_cy), AVG(part_i_summary_12_total_revenue_cy) FROM "{table_name}" WHERE part_i_summary_12_total_revenue_cy > 0'
                )
                income_stats = cursor.fetchone()

            cursor.execute(f'SELECT COUNT(DISTINCT st) FROM "{table_name}"')
            state_count = cursor.fetchone()[0]

        return {
            "success": True,
            "dataset": dataset,
            "total_records": total_count,
            "income_stats": {
                "min": income_stats[0],
                "max": income_stats[1],
                "avg": income_stats[2],
            },
            "states_count": state_count,
            "supported_formats": ["CSV", "JSON", "Excel"],
            "max_export_limit": 10000,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
